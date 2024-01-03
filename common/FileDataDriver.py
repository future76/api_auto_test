#!/usr/bin/env python
# -*- coding: UTF-8 -*-
'''
@Project ：qianxun_api 
@File    ：FileDataDriver.py
@IDE     ：PyCharm 
@Author  ：future76_lly
@Date    ：2023/12/18 13:35 
'''
import openpyxl
import yaml
# from common.EncryptDateAES import ENCRYPTAES
from config import *


class FileReader:
    """
    专门用来读取和写入yaml、excel文件
    """

    # 读取excel--openpyxl -- 文件格式：.xlsx
    @staticmethod  # 直接通过类名进行调用
    def read_excel(file_path=CASEDATAURL, sheet_name=SHEETNAME):
        """
        读取Excel文件，只支持 .xlsx文件
        :param file_path: 文件路径
        :return: excel文件数据,元组的格式
        """
        # 打开现有的Excel文件或创建新的文件
        try:
            #  正常情况下直接打开
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # 选择或创建指定的工作表
        if sheet_name in workbook.sheetnames:
            # 【正常】  判断有没有对应的shtttname ，有的话把对应的数据给我加载出来
            worksheet = workbook[sheet_name]
        else:
            # 没有的话，给我新建一个
            worksheet = workbook.create_sheet(sheet_name)

        # 获取列名 --- 把第2行的数据拿出来作为我们的key值
        headers = [cell.value for cell in worksheet[2]]
        # print("所有的key", headers)

        # 将数据存储为字典,并且放在我们data当中
        data = []  # 所有的数据

        # 把小的数据从第三行开始
        for row in worksheet.iter_rows(min_row=3, values_only=True):
            # 把所有的数据直接加进去
            # data.append(dict(zip(headers, row)))

            #  当 is_true == True才应该加进去
            new_data = dict(zip(headers, row))
            if new_data["is_true"] is True:
                data.append(new_data)
        workbook.close()
        # 所有的数据
        return data

    @staticmethod
    def writeDataToExcel(file_path=CASEDATAURL, sheet_name=SHEETNAME, row=None, column=None, value=None):
        # 打开现有的Excel文件或创建新的文件
        try:
            workbook = openpyxl.load_workbook(file_path)
        except FileNotFoundError:
            workbook = openpyxl.Workbook()

        # 选择或创建指定的工作表
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
        else:
            worksheet = workbook.create_sheet(sheet_name)

        # 写入数据到指定行和列
        worksheet.cell(row=row, column=column).value = value

        # 保存修改后的文件--- 所以执行过程当中excel是要关闭的状态
        workbook.save(file_path)

    @staticmethod
    def read_yaml(file_path=YAMLDATA):
        """
        读取yaml文件
        :param file_path: 文件路径
        :return: yaml文件数据
        """
        with open(file_path, 'r', encoding="utf-8") as file:
            data = yaml.safe_load(file)
        return data

    @staticmethod
    def write_yaml(data, file_path=YAMLDATA):
        """
        写入yaml文件，写入无序没有关系，通过key获取数据
        :param data: 需要写入的数据
        :param file_path: 文件路径
        :return:
        """
        with open(file_path, 'w', encoding="utf-8") as file:
            # allow_unicode=True，避免将中文字符转换为 Unicode 转义序列
            yaml.dump(data, file, allow_unicode=True)

    # @staticmethod
    # def data_EncryptDateAES(data):
    #     newdata = {}  # 去掉前面@符号同时数据进行加密
    #     for key in data:
    #         if key[0] == "@":
    #             # 需要加密处理
    #             newdata[key[1:]] = ENCRYPTAES.encrypt(data[key])
    #         else:
    #             # 不需要加密处理
    #             newdata[key] = data[key]
    #     return newdata


if __name__ == '__main__':
    # CaseData = FileReader.read_excel()

    # 需要加密处理
    data = {"@username": "tony", "@password": "123456"}
    res = FileReader.data_EncryptDateAES(data)
    print(res)

    # 不需要加密处理
    data = {"accounts": "hami", "@pwd": "123456", "type": "username"}
    res = FileReader.data_EncryptDateAES(data)
    print(res)
